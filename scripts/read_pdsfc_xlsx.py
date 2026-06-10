import json
import sys

import openpyxl

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8")

path = r"c:\Users\SAMKWNAG\Downloads\작업실적조회_skkr_20260609.xlsx"
wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
out = {"sheets": wb.sheetnames, "rows": []}
ws = wb[wb.sheetnames[0]]
for i, row in enumerate(ws.iter_rows(max_row=6, values_only=True)):
    out["rows"].append([("" if v is None else v) for v in row])
wb.close()
out_path = r"e:\SAMKWANG AI\SAMKWANG-PROS\SAMKWANG-PROS\scripts\pdsfc_xlsx_headers.json"
with open(out_path, "w", encoding="utf-8") as f:
    json.dump(out, f, ensure_ascii=False, indent=2)
print("wrote", out_path)
